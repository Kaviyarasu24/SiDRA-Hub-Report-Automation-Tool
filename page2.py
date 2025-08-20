import pandas as pd
import os
from pathlib import Path
import openpyxl
from PIL import Image
import io

def extract_images_from_excel(excel_file):
    # Create images directory if it doesn't exist
    if not os.path.exists('images'):
        os.makedirs('images')
    
    # Load the workbook with data_only=False to get images
    wb = openpyxl.load_workbook(excel_file, data_only=False)
    sheet = wb.active
    
    current_image_path = None
    old_image_path = None
    
    # Find column indices
    ndvi_col = None
    old_ndvi_col = None
    for col in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=col).value
        if header == 'NDVI Image date':
            ndvi_col = col
        elif header == 'Old NDVI Image date':
            old_ndvi_col = col
    
    try:
        # Get the drawing objects from the sheet
        for image in sheet._images:
            # Get the column where the image is located
            col = image.anchor._from.col + 1
            
            # Save based on which column the image is in
            img_data = io.BytesIO(image._data())
            img = Image.open(img_data)
            
            if col == ndvi_col:
                current_image_path = 'images/current_ndvi.png'
                img.save(current_image_path)
                print(f"Saved current NDVI image to {current_image_path}")
            elif col == old_ndvi_col:
                old_image_path = 'images/old_ndvi.png'
                img.save(old_image_path)
                print(f"Saved old NDVI image to {old_image_path}")
                
    except Exception as e:
        print(f"Error extracting images: {e}")
    
    return current_image_path, old_image_path

def generate_page2(excel_file, template_file, output_file, current_image=None, old_image=None, field_data=None):
    # Extract images from Excel if not provided
    if current_image is None or old_image is None:
        current_image_path, old_image_path = extract_images_from_excel(excel_file)
    else:
        current_image_path, old_image_path = current_image, old_image
        print(f"Using provided image paths: {current_image_path}, {old_image_path}")
    
    # Read the Excel file for other data
    if field_data is None:
        df = pd.read_excel(excel_file)
    else:
        df = field_data
    
    # Read the HTML template
    with open(template_file, 'r', encoding='utf-8') as f:
        html_content = f.read()
    
    # Get values from Excel using the correct column names
    try:
        old_ndvi_value = str(df['Old NDVI value'].iloc[0])
    except Exception as e:
        print(f"Error getting Old NDVI value: {e}")
        old_ndvi_value = "N/A"
        
    try:
        current_ndvi_value = str(df['NDVI value'].iloc[0])
    except Exception as e:
        print(f"Error getting NDVI value: {e}")
        current_ndvi_value = "N/A"
        
    try:
        ndvi_advisory = str(df['NDVI ADVISORY'].iloc[0])
    except Exception as e:
        print(f"Error getting NDVI ADVISORY: {e}")
        ndvi_advisory = "N/A"
    
    # Get dates from Excel
    try:
        # Try to get date from Old NDVI Image date column
        old_date = df['Old NDVI Image date'].iloc[0]
        # Check if it's NaN and fall back to Old Date if needed
        import numpy as np
        if pd.isna(old_date) or old_date is None:
            print("Old NDVI Image date is NaN, using Old Date instead")
            old_date = df['Old Date'].iloc[0]
            
        if isinstance(old_date, str):
            old_date = old_date.strip()  # Remove any whitespace or newlines
            
        old_image_date = pd.to_datetime(old_date).strftime('%d/%m/%Y')
    except Exception as e:
        print(f"Error processing old date: {e}")
        old_image_date = "N/A"
        
    try:
        # Try to get date from NDVI Image date column
        current_date = df['NDVI Image date'].iloc[0]
        # Check if it's NaN and fall back to NDMI Image date if needed
        if pd.isna(current_date) or current_date is None:
            print("NDVI Image date is NaN, using NDMI Image date instead")
            current_date = df['NDMI Image date'].iloc[0]
            
            # If still NaN, try Current image column
            if pd.isna(current_date) or current_date is None:
                print("NDMI Image date is also NaN, using Current image instead")
                current_date = df['Current  image'].iloc[0]
        
        if isinstance(current_date, str):
            current_date = current_date.strip()  # Remove any whitespace or newlines
            
        new_image_date = pd.to_datetime(current_date).strftime('%d/%m/%Y')
    except Exception as e:
        print(f"Error processing current date: {e}")
        new_image_date = "N/A"
    
    print("\nDates from Excel:")
    print(f"Old Date: {old_image_date}")
    print(f"New Date: {new_image_date}")
    
    # Replace the date placeholders in the HTML
    html_content = html_content.replace('OLD IMAGE DATE:<br/>IMAGE DATE1', f'OLD IMAGE DATE:<br/>{old_image_date}')
    html_content = html_content.replace('NEW IMAGE DATE<br/>IMAGE DATE2', f'NEW IMAGE DATE<br/>{new_image_date}')
    
    # Update the image sources in the HTML
    if old_image_path:  # First box should have the old NDVI image
        html_content = html_content.replace(
            '<img alt="Old NDVI" class="w-[220px] h-[220px] object-cover" height="220" src=" " width="220"/>',
            f'<img alt="Old NDVI" class="w-[220px] h-[220px] object-cover" height="220" src="{old_image_path}" width="220"/>'
        )
    
    if current_image_path:  # Second box should have the current NDVI image
        html_content = html_content.replace(
            '<img alt="Current NDVI" class="w-[220px] h-[220px] object-cover" height="220" src=" " width="220"/>',
            f'<img alt="Current NDVI" class="w-[220px] h-[220px] object-cover" height="220" src="{current_image_path}" width="220"/>'
        )
    
    # Fix farmland.png path
    html_content = html_content.replace(
        'src="/assest/farmland.png"',
        'src="../assest/farmland.png"'
    )
    
    # Update NDVI values and advisory
    html_content = html_content.replace('OLD NDVI VALUE', old_ndvi_value)
    html_content = html_content.replace('NDVI VALUE', current_ndvi_value)
    html_content = html_content.replace('NDVI ADVISORY', ndvi_advisory)
    
    # Remove any remaining "Value: ... (Change: ...)" text if it exists
    value_change_pattern = r'<p class="mt-2">\s*Value:[^<]*<span class="font-bold">\s*[^<]*</span>\s*<span>\s*\(Change:\s*</span>\s*<span class="font-bold">\s*[^<]*</span>\s*<span>\s*\)\s*</span>\s*</p>'
    import re
    html_content = re.sub(value_change_pattern, '', html_content)


    # Save the generated HTML
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    print(f"Page 2 report generated successfully: {output_file}")

if __name__ == "__main__":
    # File paths
    excel_file = "demo.xlsx"
    template_file = "templete/page2.html"
    output_file = "output_page2.html"
    
    # Generate the report
    generate_page2(excel_file, template_file, output_file)
