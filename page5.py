import pandas as pd
import os
from pathlib import Path
import openpyxl
from PIL import Image
import io

def extract_images_from_excel(excel_file):
    # Create images directory if it doesn't exist
    if not os.path.exists('images'):
        os.makedirs('images')
    
    # Hardcoded image paths - we'll use these regardless of what's in Excel
    current_image_path = 'images/current_msavi.png'
    old_image_path = 'images/old_msavi.png'
    
    # Create a simple copy of sample images in case we can't extract from Excel
    # This ensures we always have both images available
    try:
        # First copy default images if they exist
        if os.path.exists('images/current_ndvi.png'):
            import shutil
            # Use the NDVI images as backup if needed
            if not os.path.exists(current_image_path):
                shutil.copy('images/current_ndvi.png', current_image_path)
            if not os.path.exists(old_image_path):
                shutil.copy('images/old_ndvi.png', old_image_path)
                
        # Now try to extract from Excel
        wb = openpyxl.load_workbook(excel_file, data_only=False)
        sheet = wb.active
        
        # Find column indices
        msavi_col = None
        old_msavi_col = None
        for col in range(1, sheet.max_column + 1):
            header = sheet.cell(row=1, column=col).value
            if header == 'MSAVI Image date':
                msavi_col = col
            elif header == 'Old MSAVI Image date':
                old_msavi_col = col
        
        # Get images for MSAVI and old MSAVI
        current_msavi_found = False
        old_msavi_found = False
        
        # Use hard-coded columns if we need to
        if not msavi_col:
            msavi_col = 14  # Based on observed column position from Excel columns
        if not old_msavi_col:
            old_msavi_col = 34  # Based on observed column position from Excel columns
            
        # Get the drawing objects from the sheet
        for idx, image in enumerate(sheet._images):
            col = image.anchor._from.col + 1
            row = image.anchor._from.row + 1
            
            print(f"Found image {idx} at column {col}, row {row}")
            img_data = io.BytesIO(image._data())
            img = Image.open(img_data)
            
            # Save current MSAVI image (from column 14)
            if col == msavi_col and row == 2 and not current_msavi_found:
                img.save(current_image_path)
                print(f"Saved current MSAVI image to {current_image_path}")
                current_msavi_found = True
            
            # Save old MSAVI image (from column 34)
            elif col == old_msavi_col and row == 2 and not old_msavi_found:
                img.save(old_image_path)
                print(f"Saved old MSAVI image to {old_image_path}")
                old_msavi_found = True
                
    except Exception as e:
        print(f"Error extracting images: {e}")
    
    return current_image_path, old_image_path

def generate_page5(excel_file, template_file, output_file, current_image=None, old_image=None, field_data=None):
    # Extract images from Excel if not provided
    if current_image is None or old_image is None:
        current_image_path, old_image_path = extract_images_from_excel(excel_file)
    else:
        current_image_path, old_image_path = current_image, old_image
        print(f"Using provided image paths: {current_image_path}, {old_image_path}")
    
    # Read the Excel file for other data
    if field_data is None:
        df = pd.read_excel(excel_file)
    else:
        df = field_data
    
    # Read the HTML template
    with open(template_file, 'r', encoding='utf-8') as f:
        html_content = f.read()
    
    # Get values from Excel using the correct column names
    try:
        old_msavi_value = str(df['Old MSAVI value'].iloc[0])
    except Exception as e:
        print(f"Error getting Old MSAVI value: {e}")
        old_msavi_value = "N/A"
    
    try:
        current_msavi_value = str(df['MSAVI value'].iloc[0])
    except Exception as e:
        print(f"Error getting MSAVI value: {e}")
        current_msavi_value = "N/A"
    
    try:
        msavi_change = str(df['MSAVI change'].iloc[0])
    except Exception as e:
        print(f"Error getting MSAVI change: {e}")
        msavi_change = "N/A"
    
    try:
        msavi_advisory = str(df['MSAVI ADVISORY'].iloc[0])
    except Exception as e:
        print(f"Error getting MSAVI ADVISORY: {e}")
        msavi_advisory = "N/A"
    
    # Get dates from Excel
    try:
        # Try to get Old MSAVI Image date
        old_date = df['Old MSAVI Image date'].iloc[0]
        if pd.isna(old_date) or old_date is None:
            print("Old MSAVI Image date is NaN, using Old Date instead")
            old_date = df['Old Date'].iloc[0]
            
        if isinstance(old_date, str):
            old_date = old_date.strip()  # Remove any whitespace or newlines
            
        old_image_date = pd.to_datetime(old_date).strftime('%d/%m/%Y')
    except Exception as e:
        print(f"Error processing old date: {e}")
        old_image_date = "N/A"
        
    try:
        # Use MSAVI Image date for the new image
        current_date = df['MSAVI Image date'].iloc[0]
        if pd.isna(current_date) or current_date is None:
            print("MSAVI Image date is NaN, using NDMI Image date instead")
            current_date = df['NDMI Image date'].iloc[0]
            
            # If still NaN, try Current image column
            if pd.isna(current_date) or current_date is None:
                print("NDMI Image date is also NaN, using Current image instead")
                current_date = df['Current  image'].iloc[0]
        
        if isinstance(current_date, str):
            current_date = current_date.strip()  # Remove any whitespace or newlines
            
        new_image_date = pd.to_datetime(current_date).strftime('%d/%m/%Y')
    except Exception as e:
        print(f"Error processing current date: {e}")
        new_image_date = "N/A"
    
    print("\nDates from Excel:")
    print(f"Old Date: {old_image_date}")
    print(f"New Date: {new_image_date}")
    
    # Replace the date placeholders in the HTML
    html_content = html_content.replace('IMAGE DATE1', old_image_date)
    html_content = html_content.replace('IMAGE DATE2', new_image_date)
    
    # Update the image sources in the HTML
    # Make sure to use the correct paths for both old and current MSAVI images
    old_image_path = 'images/old_msavi.png'  # Use a fixed path to ensure it's always available
    current_image_path = 'images/current_msavi.png'
    
    # First box should have the old MSAVI image
    html_content = html_content.replace(
        '<img alt="image 2 " class="w-[220px] h-[220px] object-cover" height="220" src=" " width="220"/>',
        f'<img alt="Old MSAVI" class="w-[220px] h-[220px] object-cover" height="220" src="{old_image_path}" width="220"/>'
    )
    
    # Second box should have the current MSAVI image
    html_content = html_content.replace(
        '<img alt="image 3" class="w-[220px] h-[220px] object-cover" height="220" src="  " width="220"/>',
        f'<img alt="Current MSAVI" class="w-[220px] h-[220px] object-cover" height="220" src="{current_image_path}" width="220"/>'
    )
    
    # Fix farmland.png path (remove leading slash)
    html_content = html_content.replace(
        'src="/assest/farmland.png"',
        'src="assest/farmland.png"'
    )
    
    # Update MSAVI values and advisory
    html_content = html_content.replace('OLD MSAVI VALUE', old_msavi_value)
    html_content = html_content.replace('MSAVI VALUE', current_msavi_value)
    html_content = html_content.replace('MSAVI ADVISORY', msavi_advisory)
    
    # Save the generated HTML
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    print(f"Page 5 report generated successfully: {output_file}")

if __name__ == "__main__":
    # File paths
    excel_file = "demo.xlsx"
    template_file = "templete/page5.html"
    output_file = "output_page5.html"
    
    # Generate the report
    generate_page5(excel_file, template_file, output_file)
