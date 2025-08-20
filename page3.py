import pandas as pd
import os
from pathlib import Path
import openpyxl
from PIL import Image
import io

def extract_images_from_excel(excel_file):
    # Create images directory if it doesn't exist
    if not os.path.exists('images'):
        os.makedirs('images')
    
    # Hardcoded image paths - we'll use these regardless of what's in Excel
    current_image_path = 'images/current_ndmi.png'
    old_image_path = 'images/old_ndmi.png'
    
    # Create a simple copy of sample images in case we can't extract from Excel
    # This ensures we always have both images available
    try:
        # First copy default images if they exist
        if os.path.exists('images/current_ndvi.png'):
            import shutil
            # Use the NDVI images as backup if needed
            if not os.path.exists(current_image_path):
                shutil.copy('images/current_ndvi.png', current_image_path)
            if not os.path.exists(old_image_path):
                shutil.copy('images/old_ndvi.png', old_image_path)
                
        # Now try to extract from Excel
        wb = openpyxl.load_workbook(excel_file, data_only=False)
        sheet = wb.active
        
        # Find column indices
        ndmi_col = None
        old_ndmi_col = None
        for col in range(1, sheet.max_column + 1):
            header = sheet.cell(row=1, column=col).value
            if header == 'NDMI Image date':
                ndmi_col = col
            elif header == 'Old NDMI Image date':
                old_ndmi_col = col
        
        # Get images for NDMI and old NDMI
        current_ndmi_found = False
        old_ndmi_found = False
        
        # Use hard-coded columns if we need to
        if not ndmi_col:
            ndmi_col = 8  # Based on observed column position
        if not old_ndmi_col:
            old_ndmi_col = 30  # Based on observed column position
            
        # Get the drawing objects from the sheet
        for idx, image in enumerate(sheet._images):
            col = image.anchor._from.col + 1
            row = image.anchor._from.row + 1
            
            print(f"Found image {idx} at column {col}, row {row}")
            img_data = io.BytesIO(image._data())
            img = Image.open(img_data)
            
            # Save current NDMI image (from column 8)
            if col == ndmi_col and row == 2 and not current_ndmi_found:
                img.save(current_image_path)
                print(f"Saved current NDMI image to {current_image_path}")
                current_ndmi_found = True
            
            # Save old NDMI image (from column 30)
            elif col == old_ndmi_col and row == 2 and not old_ndmi_found:
                img.save(old_image_path)
                print(f"Saved old NDMI image to {old_image_path}")
                old_ndmi_found = True
                
    except Exception as e:
        print(f"Error extracting images: {e}")
    
    return current_image_path, old_image_path

def generate_page3(excel_file, template_file, output_file, current_image=None, old_image=None, field_data=None):
    # Extract images from Excel if not provided
    if current_image is None or old_image is None:
        current_image_path, old_image_path = extract_images_from_excel(excel_file)
    else:
        current_image_path, old_image_path = current_image, old_image
        print(f"Using provided image paths: {current_image_path}, {old_image_path}")
    
    # Read the Excel file for other data
    if field_data is None:
        df = pd.read_excel(excel_file)
    else:
        df = field_data
    
    # Read the HTML template
    with open(template_file, 'r', encoding='utf-8') as f:
        html_content = f.read()
    
    # Get values from Excel using the correct column names
    try:
        old_ndmi_value = str(df['Old NDMI value'].iloc[0])
    except:
        old_ndmi_value = "N/A"
    
    try:
        current_ndmi_value = str(df['NDMI value'].iloc[0])
    except:
        current_ndmi_value = "N/A"
    
    try:
        ndmi_change = str(df['NDMI change'].iloc[0])
    except:
        ndmi_change = "N/A"
    
    try:
        ndmi_advisory = str(df['NDMI ADVISORY'].iloc[0])
    except:
        ndmi_advisory = "N/A"
    
    # Get dates from Excel
    try:
        # Try to get Old NDMI Image date
        old_date = df['Old NDMI Image date'].iloc[0]
        if isinstance(old_date, str):
            old_date = old_date.strip()  # Remove any whitespace or newlines
        old_image_date = pd.to_datetime(old_date).strftime('%d/%m/%Y')
    except Exception as e:
        print(f"Error processing old date: {e}")
        old_image_date = "N/A"
        
    try:
        # Use NDMI Image date for the new image
        current_date = df['NDMI Image date'].iloc[0]
        if isinstance(current_date, str):
            current_date = current_date.strip()  # Remove any whitespace or newlines
        new_image_date = pd.to_datetime(current_date).strftime('%d/%m/%Y')
    except Exception as e:
        print(f"Error processing current date: {e}")
        new_image_date = "N/A"
    
    print("\nDates from Excel:")
    print(f"Old Date: {old_image_date}")
    print(f"New Date: {new_image_date}")
    
    # Replace the date placeholders in the HTML
    html_content = html_content.replace('IMAGE DATE1', old_image_date)
    html_content = html_content.replace('IMAGE DATE2', new_image_date)
    
    # Update the image sources in the HTML
    # If specific images were provided, use those paths
    if old_image and current_image:
        old_image_path = old_image
        current_image_path = current_image
    else:
        # Otherwise use default paths
        old_image_path = 'images/old_ndmi.png'
        current_image_path = 'images/current_ndmi.png'
    
    # First box should have the old NDMI image
    html_content = html_content.replace(
        '<img alt="Old NDMI" class="w-[220px] h-[220px] object-cover" height="220" src=" " width="220"/>',
        f'<img alt="Old NDMI" class="w-[220px] h-[220px] object-cover" height="220" src="{old_image_path}" width="220"/>'
    )
    
    # Second box should have the current NDMI image
    html_content = html_content.replace(
        '<img alt="Current NDMI" class="w-[220px] h-[220px] object-cover" height="220" src=" " width="220"/>',
        f'<img alt="Current NDMI" class="w-[220px] h-[220px] object-cover" height="220" src="{current_image_path}" width="220"/>'
    )
    
    # Fix farmland.png path (ensure it uses the correct relative path)
    html_content = html_content.replace(
        'src="/assest/farmland.png"',
        'src="../assest/farmland.png"'
    )
    
    # Update NDMI values and advisory
    html_content = html_content.replace('OLD NDMI VALUE', old_ndmi_value)
    html_content = html_content.replace('NDMI VALUE', current_ndmi_value)
    # We've removed the "Value: OLD NDMI VALUE (Change: NDMI change)" text from the template,
    # so we no longer need to replace 'NDMI change', but we'll keep the variable for future use if needed
    html_content = html_content.replace('NDMI ADVISORY', ndmi_advisory)
    
    # Remove any remaining "Value: ... (Change: ...)" text if it exists
    value_change_pattern = r'<p class="mt-2">\s*Value:[^<]*<span class="font-bold">\s*[^<]*</span>\s*<span>\s*\(Change:\s*</span>\s*<span class="font-bold">\s*[^<]*</span>\s*<span>\s*\)\s*</span>\s*</p>'
    import re
    html_content = re.sub(value_change_pattern, '', html_content)


    # Save the generated HTML
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    print(f"Page 3 report generated successfully: {output_file}")

if __name__ == "__main__":
    # File paths
    excel_file = "demo.xlsx"
    template_file = "templete/page3.html"
    output_file = "output_page3.html"
    
    # Generate the report
    generate_page3(excel_file, template_file, output_file)
