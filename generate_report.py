import os
import pandas as pd
import page1
import page2
import page3
import page4
import page5
import page6
import openpyxl
from PIL import Image
import io
import shutil

def extract_field_images(excel_file, row, output_dir):
    """
    Extract images for a specific field from the Excel file
    
    Args:
        excel_file (str): Path to the Excel file with crop data
        row (pandas.Series): The row data for the field
        output_dir (str): Directory where images will be saved
    """
    # Load the workbook
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=False)
        sheet = wb.active
        
        # Use field-specific folder to save images
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        # Initialize default images (copy from existing if available)
        default_image_pairs = [
            ("current_ndvi.png", "old_ndvi.png"),
            ("current_ndmi.png", "old_ndmi.png"),
            ("current_reci.png", "old_reci.png"),
            ("current_msavi.png", "old_msavi.png"),
            ("current_ndre.png", "old_ndre.png")
        ]
        
        # Create default copies from existing images if available
        for current_img, old_img in default_image_pairs:
            src_current = os.path.join("images", current_img)
            src_old = os.path.join("images", old_img)
            
            dest_current = os.path.join(output_dir, current_img)
            dest_old = os.path.join(output_dir, old_img)
            
            # Copy default images if they exist
            if os.path.exists(src_current) and not os.path.exists(dest_current):
                shutil.copy(src_current, dest_current)
            
            if os.path.exists(src_old) and not os.path.exists(dest_old):
                shutil.copy(src_old, dest_old)
        
        # Now extract the field-specific images
        # Find the row in Excel that matches this field
        field_name = row['Field'] if 'Field' in row else None
        excel_row = None
        
        for r in range(1, sheet.max_row + 1):
            if sheet.cell(row=r, column=1).value == field_name:
                excel_row = r
                break
        
        if excel_row is None:
            print(f"Could not find row for field {field_name} in Excel")
            return
        
        print(f"Found field {field_name} at row {excel_row} in Excel")
        
        # Column indices for different image types
        column_indices = {
            'NDVI Image date': ('current_ndvi.png', None),
            'Old NDVI Image date': ('old_ndvi.png', None),
            'NDMI Image date': ('current_ndmi.png', None),
            'Old NDMI Image date': ('old_ndmi.png', None),
            'RECI Image date': ('current_reci.png', None),
            'Old RECI Image date': ('old_reci.png', None),
            'MSAVI Image date': ('current_msavi.png', None),
            'Old MSAVI Image date': ('old_msavi.png', None),
            'NDRE Image date': ('current_ndre.png', None),
            'Old NDRE Image date': ('old_ndre.png', None)
        }
        
        # Find column indices
        for col in range(1, sheet.max_column + 1):
            header = sheet.cell(row=1, column=col).value
            if header in column_indices:
                image_file, _ = column_indices[header]
                column_indices[header] = (image_file, col)
        
        # Extract images from the specific row
        for image in sheet._images:
            col = image.anchor._from.col + 1
            img_row = image.anchor._from.row + 1
            
            # Only process images in the row for this field
            if img_row != excel_row:
                continue
                
            # Find which image type this column corresponds to
            for header, (image_file, col_idx) in column_indices.items():
                if col_idx == col:
                    # Extract and save this image
                    img_data = io.BytesIO(image._data())
                    img = Image.open(img_data)
                    output_path = os.path.join(output_dir, image_file)
                    img.save(output_path)
                    print(f"Saved {header} image for {field_name} to {output_path}")
                    break
    
    except Exception as e:
        print(f"Error extracting field images: {e}")

def generate_full_report(excel_file, output_directory="reports"):
    """
    Generate a comprehensive report with all pages for each field in the Excel file
    
    Args:
        excel_file (str): Path to the Excel file with crop data
        output_directory (str): Directory where the reports will be saved
    """
    # Create output directory if it doesn't exist
    if not os.path.exists(output_directory):
        os.makedirs(output_directory)
    
    # Read Excel data
    try:
        df = pd.read_excel(excel_file)
        print(f"Successfully read Excel file: {excel_file}")
        print(f"Found {len(df)} rows of data")
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return
    
    # Create folder for field-specific images
    images_dir = "images"
    if not os.path.exists(images_dir):
        os.makedirs(images_dir)
    
    # Process each row and generate individual reports
    for index, row in df.iterrows():
        # Get field name for the report filename
        field_name = str(row['Field']).replace(' ', '_').replace('/', '_') if 'Field' in row else f"field_{index+1}"
        print(f"\n===== Generating report for {field_name} =====")
        
        # Create field-specific image folder for this report
        field_images_dir = os.path.join(images_dir, field_name)
        if not os.path.exists(field_images_dir):
            os.makedirs(field_images_dir)
        
        # Create temporary files for each page
        temp_files = {
            "page1": "temp_page1.html",
            "page2": "temp_page2.html",
            "page3": "temp_page3.html",
            "page4": "temp_page4.html",
            "page5": "temp_page5.html",
            "page6": "temp_page6.html"
        }
        
        # Create single row dataframe with this row
        single_row_data = pd.DataFrame([row])
        
        try:
            # Extract row-specific images from Excel first
            extract_field_images(excel_file, row, field_images_dir)
            
            # Page 1 - Field Information
            print("Generating Page 1: Field Information")
            page1.generate_report_html(single_row_data, "templete/page1.html", temp_files["page1"])
            
            # Page 2 - NDVI (Green Health Score)
            print("Generating Page 2: NDVI (Green Health Score)")
            # Since we already extracted the images for this field, override the image paths
            page2.generate_page2(excel_file, "templete/page2.html", temp_files["page2"], 
                                current_image=os.path.join(field_images_dir, "current_ndvi.png"),
                                old_image=os.path.join(field_images_dir, "old_ndvi.png"),
                                field_data=single_row_data)
            
            # Page 3 - NDMI (Moisture Level Indicator) 
            print("Generating Page 3: NDMI (Moisture Level Indicator)")
            page3.generate_page3(excel_file, "templete/page3.html", temp_files["page3"],
                               current_image=os.path.join(field_images_dir, "current_ndmi.png"),
                               old_image=os.path.join(field_images_dir, "old_ndmi.png"),
                               field_data=single_row_data)
            
            # Page 4 - RECI (Leaf Freshness Index)
            print("Generating Page 4: RECI (Leaf Freshness Index)")
            page4.generate_page4(excel_file, "templete/page4.html", temp_files["page4"],
                               current_image=os.path.join(field_images_dir, "current_reci.png"),
                               old_image=os.path.join(field_images_dir, "old_reci.png"),
                               field_data=single_row_data)
            
            # Page 5 - MSAVI (Growth Strength Index)
            print("Generating Page 5: MSAVI (Growth Strength Index)")
            page5.generate_page5(excel_file, "templete/page5.html", temp_files["page5"],
                               current_image=os.path.join(field_images_dir, "current_msavi.png"),
                               old_image=os.path.join(field_images_dir, "old_msavi.png"),
                               field_data=single_row_data)
            
            # Page 6 - NDRE (Early Stress Checker)
            print("Generating Page 6: NDRE (Early Stress Checker)")
            page6.generate_page6(excel_file, "templete/page6.html", temp_files["page6"],
                               current_image=os.path.join(field_images_dir, "current_ndre.png"),
                               old_image=os.path.join(field_images_dir, "old_ndre.png"),
                               field_data=single_row_data)
            
            # Combine all pages into one report
            combined_html = combine_html_pages(temp_files, field_name)
            
            # Save the combined report
            output_path = os.path.join(output_directory, f"full_report_{field_name}.html")
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(combined_html)
                
            print(f"Full report generated successfully: {output_path}")
            
        except Exception as e:
            print(f"Error generating report for {field_name}: {e}")
        finally:
            # Clean up temporary files
            for temp_file in temp_files.values():
                if os.path.exists(temp_file):
                    try:
                        os.remove(temp_file)
                    except Exception:
                        pass

def combine_html_pages(page_files, field_name=""):
    """
    Combine multiple HTML pages into a single HTML document
    
    Args:
        page_files (dict): Dictionary of page names and their file paths
        field_name (str): Name of the field for this report
        
    Returns:
        str: Combined HTML content
    """
    # Create output filename from field name
    page_name = f"full_report_{field_name}"
    # Start with a basic HTML structure
    combined_html = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="utf-8"/>
    <meta content="width=device-width, initial-scale=1" name="viewport"/>
    <title>SiDRA Hub Crop Report</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/5.15.3/css/all.min.css" rel="stylesheet"/>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/html2pdf.js/0.10.1/html2pdf.bundle.min.js"></script>
    <style>
        @media print {
            .page-break {
                page-break-after: always;
            }
        }
        .page {
            margin-bottom: 40px;
        }
    </style>
</head>
<body class="bg-gray-100">
    <!-- PDF Download Button -->
    <div class="fixed top-4 right-4 z-50">
        <button id="downloadPdf" class="bg-blue-600 hover:bg-blue-700 text-white font-bold py-2 px-4 rounded-lg shadow-lg flex items-center gap-2 transition-colors">
            <i class="fas fa-download"></i>
            Download PDF
        </button>
    </div>
    
    <div id="reportContent">
"""
    
    # Read and combine each page's content
    for page_name, file_path in page_files.items():
        try:
            if os.path.exists(file_path):
                with open(file_path, 'r', encoding='utf-8') as f:
                    html_content = f.read()
                
                # Extract the body content between <body> and </body>
                start_idx = html_content.find("<body")
                if start_idx != -1:
                    start_idx = html_content.find(">", start_idx) + 1
                    end_idx = html_content.find("</body>", start_idx)
                    if start_idx != -1 and end_idx != -1:
                        body_content = html_content[start_idx:end_idx].strip()
                        
                        # Remove any existing PDF download buttons or scripts
                        download_btn_idx = body_content.find('id="downloadPdf"')
                        if download_btn_idx != -1:
                            # Find the surrounding div
                            div_start = body_content.rfind('<div', 0, download_btn_idx)
                            div_end = body_content.find('</div>', download_btn_idx)
                            if div_start != -1 and div_end != -1:
                                body_content = body_content[:div_start] + body_content[div_end+6:]
                        
                        # Clean up any PDF generation scripts
                        script_idx = body_content.find('<script')
                        while script_idx != -1:
                            script_end = body_content.find('</script>', script_idx)
                            if script_end != -1:
                                body_content = body_content[:script_idx] + body_content[script_end+9:]
                                script_idx = body_content.find('<script')
                            else:
                                break
                        
                        # Fix image paths before adding to combined HTML
                        # Replace direct image paths to point to correct location from reports directory
                        body_content = body_content.replace('src="images\\', 'src="../images/')
                        body_content = body_content.replace('src="images/', 'src="../images/')
                        body_content = body_content.replace('src="assest/', 'src="../assest/')
                        
                        # Make sure double replacements don't happen
                        body_content = body_content.replace('src="../../assest/', 'src="../assest/')
                        body_content = body_content.replace('src="../../images/', 'src="../images/')
                        
                        # Direct approach to fix corrupted HTML
                        import re
                        
                        # Use the field_name passed to the function
                        field_path = field_name
                                                
                        print(f"Using field path for images: {field_path}")
                        
                        # First, manually fix the malformed image tags
                        # Look for patterns like 'src="../images/old_ndmi.png" width="220"/old_ndmi.png" width="220"'
                        body_content = re.sub(
                            r'src="\.\.\/images\/(old|current)_(ndvi|ndmi|reci|msavi|ndre)\.png" width="(\d+)"\/(old|current)_(ndvi|ndmi|reci|msavi|ndre)\.png" width="\3"\/>',
                            r'src="../images/\1_\2.png" width="\3"/>',
                            body_content
                        )
                        
                        # Now replace all non-field-specific image paths with field-specific ones
                        for time_prefix in ['old', 'current']:
                            for index_type in ['ndvi', 'ndmi', 'reci', 'msavi', 'ndre']:
                                # Replace src="../images/current_ndvi.png" with src="../images/field_name/current_ndvi.png"
                                generic_path = f'src="../images/{time_prefix}_{index_type}.png"'
                                specific_path = f'src="../images/{field_path}/{time_prefix}_{index_type}.png"'
                                body_content = body_content.replace(generic_path, specific_path)                        # Wrap each page in a div with page-break
                        combined_html += f"""
        <div class="page" id="{page_name}">
            {body_content}
        </div>
        <div class="page-break"></div>
"""
        except Exception as e:
            print(f"Error processing {page_name} ({file_path}): {e}")
    
    # Add closing tags and PDF generation script
    combined_html += """
    </div>

    <script>
        document.getElementById('downloadPdf').addEventListener('click', function() {
            console.log('PDF download button clicked');
            
            // Hide the download button during PDF generation
            this.style.display = 'none';
            
            const element = document.getElementById('reportContent');
            console.log('Report element found:', element);
            
            if (!element) {
                alert('Report content not found!');
                this.style.display = 'flex';
                return;
            }
            
            // Wait for any dynamic content to load
            setTimeout(() => {
                console.log('Starting PDF generation...');
                
                // Optimized options for better PDF quality and reliability
                const opt = {
                    margin: 10,
                    filename: 'sidra_crop_report.pdf',
                    image: { 
                        type: 'jpeg', 
                        quality: 0.98
                    },
                    html2canvas: { 
                        scale: 2,
                        useCORS: true,
                        allowTaint: true,
                        letterRendering: true,
                        backgroundColor: '#dbe8f2'
                    },
                    jsPDF: { 
                        unit: 'mm', 
                        format: 'a4', 
                        orientation: 'portrait'
                    },
                    pagebreak: { mode: ['avoid-all', 'css', 'legacy'] }
                };
                
                // Generate PDF
                html2pdf()
                    .from(element)
                    .set(opt)
                    .save()
                    .then(() => {
                        console.log('PDF generation successful');
                        document.getElementById('downloadPdf').style.display = 'flex';
                    })
                    .catch((error) => {
                        console.error('PDF generation failed:', error);
                        document.getElementById('downloadPdf').style.display = 'flex';
                        alert('PDF generation failed: ' + error.message);
                    });
            }, 500);
        });
    </script>
</body>
</html>
"""
    
    return combined_html

if __name__ == "__main__":
    # File paths
    excel_file = "demo.xlsx"
    output_directory = "reports"
    
    # Generate the full report
    generate_full_report(excel_file, output_directory)
