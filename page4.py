import pandas as pd
import os
from pathlib import Path
import openpyxl
from PIL import Image
import io

def extract_images_from_excel(excel_file):
    # Create images directory if it doesn't exist
    if not os.path.exists('images'):
        os.makedirs('images')
    
    # Hardcoded image paths - we'll use these regardless of what's in Excel
    current_image_path = 'images/current_reci.png'
    old_image_path = 'images/old_reci.png'
    
    # Create a simple copy of sample images in case we can't extract from Excel
    # This ensures we always have both images available
    try:
        # First copy default images if they exist
        if os.path.exists('images/current_ndvi.png'):
            import shutil
            # Use the NDVI images as backup if needed
            if not os.path.exists(current_image_path):
                shutil.copy('images/current_ndvi.png', current_image_path)
            if not os.path.exists(old_image_path):
                shutil.copy('images/old_ndvi.png', old_image_path)
                
        # Now try to extract from Excel
        wb = openpyxl.load_workbook(excel_file, data_only=False)
        sheet = wb.active
        
        # Find column indices
        reci_col = None
        old_reci_col = None
        for col in range(1, sheet.max_column + 1):
            header = sheet.cell(row=1, column=col).value
            if header == 'RECI Image date':
                reci_col = col
            elif header == 'Old RECI Image date':
                old_reci_col = col
        
        # Get images for RECI and old RECI
        current_reci_found = False
        old_reci_found = False
        
        # Use hard-coded columns if we need to
        if not reci_col:
            reci_col = 11  # Based on observed column position from Excel columns
        if not old_reci_col:
            old_reci_col = 32  # Based on observed column position from Excel columns
            
        # Get the drawing objects from the sheet
        for idx, image in enumerate(sheet._images):
            col = image.anchor._from.col + 1
            row = image.anchor._from.row + 1
            
            print(f"Found image {idx} at column {col}, row {row}")
            img_data = io.BytesIO(image._data())
            img = Image.open(img_data)
            
            # Save current RECI image (from column 11)
            if col == reci_col and row == 2 and not current_reci_found:
                img.save(current_image_path)
                print(f"Saved current RECI image to {current_image_path}")
                current_reci_found = True
            
            # Save old RECI image (from column 32)
            elif col == old_reci_col and row == 2 and not old_reci_found:
                img.save(old_image_path)
                print(f"Saved old RECI image to {old_image_path}")
                old_reci_found = True
                
    except Exception as e:
        print(f"Error extracting images: {e}")
    
    return current_image_path, old_image_path

def generate_page4(excel_file, template_file, output_file, current_image=None, old_image=None, field_data=None):
    # Extract images from Excel if not provided
    if current_image is None or old_image is None:
        current_image_path, old_image_path = extract_images_from_excel(excel_file)
    else:
        current_image_path, old_image_path = current_image, old_image
        print(f"Using provided image paths: {current_image_path}, {old_image_path}")
    
    # Read the Excel file for other data
    if field_data is None:
        df = pd.read_excel(excel_file)
    else:
        df = field_data
    
    # Read the HTML template
    with open(template_file, 'r', encoding='utf-8') as f:
        html_content = f.read()
    
    # Get values from Excel using the correct column names
    try:
        old_reci_value = str(df['Old RECI value'].iloc[0])
    except Exception as e:
        print(f"Error getting Old RECI value: {e}")
        old_reci_value = "N/A"
    
    try:
        current_reci_value = str(df['RECI value'].iloc[0])
    except Exception as e:
        print(f"Error getting RECI value: {e}")
        current_reci_value = "N/A"
    
    try:
        reci_change = str(df['RECI change'].iloc[0])
    except Exception as e:
        print(f"Error getting RECI change: {e}")
        reci_change = "N/A"
    
    try:
        reci_advisory = str(df['RECI ADVISORY'].iloc[0])
    except Exception as e:
        print(f"Error getting RECI ADVISORY: {e}")
        reci_advisory = "N/A"
    
    # Get dates from Excel
    try:
        # Try to get Old RECI Image date
        old_date = df['Old RECI Image date'].iloc[0]
        if pd.isna(old_date) or old_date is None:
            print("Old RECI Image date is NaN, using Old Date instead")
            old_date = df['Old Date'].iloc[0]
            
        if isinstance(old_date, str):
            old_date = old_date.strip()  # Remove any whitespace or newlines
            
        old_image_date = pd.to_datetime(old_date).strftime('%d/%m/%Y')
    except Exception as e:
        print(f"Error processing old date: {e}")
        old_image_date = "N/A"
        
    try:
        # Use RECI Image date for the new image
        current_date = df['RECI Image date'].iloc[0]
        if pd.isna(current_date) or current_date is None:
            print("RECI Image date is NaN, using NDMI Image date instead")
            current_date = df['NDMI Image date'].iloc[0]
            
            # If still NaN, try Current image column
            if pd.isna(current_date) or current_date is None:
                print("NDMI Image date is also NaN, using Current image instead")
                current_date = df['Current  image'].iloc[0]
        
        if isinstance(current_date, str):
            current_date = current_date.strip()  # Remove any whitespace or newlines
            
        new_image_date = pd.to_datetime(current_date).strftime('%d/%m/%Y')
    except Exception as e:
        print(f"Error processing current date: {e}")
        new_image_date = "N/A"
    
    print("\nDates from Excel:")
    print(f"Old Date: {old_image_date}")
    print(f"New Date: {new_image_date}")
    
    # Replace the date placeholders in the HTML
    html_content = html_content.replace('IMAGE DATE1', old_image_date)
    html_content = html_content.replace('IMAGE DATE2', new_image_date)
    
    # Update the image sources in the HTML
    # If specific images were provided, use those paths
    if old_image and current_image:
        old_image_path = old_image
        current_image_path = current_image
    else:
        # Otherwise use default paths
        old_image_path = 'images/old_reci.png'
        current_image_path = 'images/current_reci.png'
    
    # First box should have the old RECI image
    html_content = html_content.replace(
        '<img alt="Old RECI" class="w-[220px] h-[220px] object-cover" height="220" src=" " width="220"/>',
        f'<img alt="Old RECI" class="w-[220px] h-[220px] object-cover" height="220" src="{old_image_path}" width="220"/>'
    )
    
    # Second box should have the current RECI image
    html_content = html_content.replace(
        '<img alt="Current RECI" class="w-[220px] h-[220px] object-cover" height="220" src=" " width="220"/>',
        f'<img alt="Current RECI" class="w-[220px] h-[220px] object-cover" height="220" src="{current_image_path}" width="220"/>'
    )
    
    # Fix farmland.png path (ensure it uses the correct relative path)
    html_content = html_content.replace(
        'src="/assest/farmland.png"',
        'src="../assest/farmland.png"'
    )
    
    # Update RECI values and advisory
    html_content = html_content.replace('OLD RECI VALUE', old_reci_value)
    html_content = html_content.replace('RECI VALUE', current_reci_value)
    html_content = html_content.replace('RECI ADVISORY', reci_advisory)
    
    # Save the generated HTML
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    print(f"Page 4 report generated successfully: {output_file}")

if __name__ == "__main__":
    # File paths
    excel_file = "demo.xlsx"
    template_file = "templete/page4.html"
    output_file = "output_page4.html"
    
    # Generate the report
    generate_page4(excel_file, template_file, output_file)
